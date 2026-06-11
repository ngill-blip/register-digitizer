"""
Form Digitizer — Flask backend
Preprocessing: OpenCV  |  OCR: Tesseract  |  Export: openpyxl / csv
"""

import os
import re
import csv
import uuid
import json
import math
import tempfile
from pathlib import Path
from io import BytesIO, StringIO

import cv2
import numpy as np
import pytesseract
from PIL import Image, ImageEnhance

from flask import (
    Flask, request, jsonify, send_file,
    render_template, send_from_directory
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from werkzeug.utils import secure_filename

# ── Config ──────────────────────────────────────────────────────────────────

BASE_DIR   = Path(__file__).parent
UPLOAD_DIR = BASE_DIR / "uploads"

# Exports saved to a fixed folder outside the app so they survive updates/reinstalls.
# Falls back to app/exports if the Desktop location isn't available.
_PERMANENT_EXPORT_DIR = Path.home() / "Desktop" / "Claude (DND)" / "Form Exports"
try:
    _PERMANENT_EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    EXPORT_DIR = _PERMANENT_EXPORT_DIR
except OSError:
    EXPORT_DIR = BASE_DIR / "exports"
    EXPORT_DIR.mkdir(exist_ok=True)

print(f"📁  Exports folder: {EXPORT_DIR}")

UPLOAD_DIR.mkdir(exist_ok=True)

ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "webp", "bmp", "tiff", "tif", "pdf"}
MAX_CONTENT_MB     = 20

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = MAX_CONTENT_MB * 1024 * 1024
app.secret_key = os.environ.get("SECRET_KEY", "dev-key-change-in-prod")


# ── Helpers ──────────────────────────────────────────────────────────────────

def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def preprocess_image(pil_img: Image.Image) -> np.ndarray:
    """
    Full preprocessing pipeline:
    1. Convert to grayscale
    2. Auto-rotate based on EXIF orientation
    3. Deskew
    4. Denoise
    5. Adaptive threshold (binarise)
    Returns a numpy uint8 array suitable for Tesseract.
    """
    # EXIF-aware rotation via PIL
    pil_img = fix_orientation(pil_img)

    # Resize if very large (speeds up OCR, keeps accuracy)
    max_dim = 3000
    w, h = pil_img.size
    if max(w, h) > max_dim:
        scale = max_dim / max(w, h)
        pil_img = pil_img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)

    img = np.array(pil_img.convert("RGB"))
    gray = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)

    # Deskew
    gray = deskew(gray)

    # Denoise
    gray = cv2.fastNlMeansDenoising(gray, h=10)

    # Enhance contrast
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    gray  = clahe.apply(gray)

    # Adaptive threshold → cleaner text
    binary = cv2.adaptiveThreshold(
        gray, 255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY, 31, 10
    )

    return binary


def fix_orientation(img: Image.Image) -> Image.Image:
    """Rotate image according to EXIF orientation tag."""
    try:
        exif = img._getexif()
        if exif:
            orientation = exif.get(274)  # tag 274 = Orientation
            rotations = {3: 180, 6: 270, 8: 90}
            if orientation in rotations:
                img = img.rotate(rotations[orientation], expand=True)
    except Exception:
        pass
    return img


def deskew(gray: np.ndarray) -> np.ndarray:
    """Detect and correct skew angle using Hough transform."""
    try:
        edges  = cv2.Canny(gray, 50, 150, apertureSize=3)
        lines  = cv2.HoughLinesP(edges, 1, np.pi / 180, 100,
                                 minLineLength=100, maxLineGap=10)
        if lines is None:
            return gray
        angles = []
        for line in lines:
            x1, y1, x2, y2 = line[0]
            if x2 != x1:
                angle = math.degrees(math.atan2(y2 - y1, x2 - x1))
                if -45 < angle < 45:
                    angles.append(angle)
        if not angles:
            return gray
        median_angle = float(np.median(angles))
        if abs(median_angle) < 0.5:
            return gray
        h, w = gray.shape
        M   = cv2.getRotationMatrix2D((w / 2, h / 2), median_angle, 1)
        out = cv2.warpAffine(gray, M, (w, h),
                             flags=cv2.INTER_CUBIC,
                             borderMode=cv2.BORDER_REPLICATE)
        return out
    except Exception:
        return gray


# ── Language detection map ────────────────────────────────────────────────────
# Maps langdetect ISO codes → Tesseract language pack + display name
LANG_MAP = {
    "sw":  ("eng+swa", "Swahili"),
    "sn":  ("eng+sna", "Shona"),
    "af":  ("eng+afr", "Afrikaans"),
    "zu":  ("eng+zul", "Zulu"),
    "xh":  ("eng+xho", "Xhosa"),
    "nd":  ("eng+nde", "Ndebele"),
    "hi":  ("eng+hin", "Hindi"),
    "ur":  ("eng+urd", "Urdu"),
    "fr":  ("eng+fra", "French"),
    "pt":  ("eng+por", "Portuguese"),
    "en":  ("eng",     "English"),
}

def detect_language(text: str) -> tuple[str, str]:
    """
    Detect language from OCR text.
    Returns (tesseract_lang_string, display_name).
    Falls back to English if detection fails or lang is unsupported.
    """
    try:
        from langdetect import detect, LangDetectException
        if len(text.strip()) < 20:
            return "eng", "English"
        detected = detect(text)
        return LANG_MAP.get(detected, ("eng", f"English (detected: {detected})"))
    except Exception:
        return "eng", "English"


def run_ocr(preprocessed: np.ndarray, lang: str = "eng") -> dict:
    """
    Two-pass OCR with automatic language detection.
    Pass 1: Quick English OCR to get enough text for language detection.
    Pass 2: Re-run with the detected language pack for best accuracy.
    Returns dict with full_text, word-level bounding boxes, and detected language.
    """
    custom_config = "--oem 3 --psm 3"

    # Pass 1 — English only (fast)
    try:
        first_pass = pytesseract.image_to_string(
            preprocessed, lang="eng", config=custom_config
        )
    except Exception:
        first_pass = ""

    # Detect language from first pass text
    detected_lang, detected_name = detect_language(first_pass)

    # Pass 2 — re-run with detected language if different from English
    final_lang = detected_lang if detected_lang != lang else lang
    try:
        data = pytesseract.image_to_data(
            preprocessed, lang=final_lang, config=custom_config,
            output_type=pytesseract.Output.DICT
        )
        full_text = pytesseract.image_to_string(
            preprocessed, lang=final_lang, config=custom_config
        )
    except Exception:
        # Language pack not installed — fall back to first pass result
        data = pytesseract.image_to_data(
            preprocessed, lang="eng", config=custom_config,
            output_type=pytesseract.Output.DICT
        )
        full_text = first_pass
        detected_name = "English (fallback)"

    return {"full_text": full_text, "data": data, "detected_language": detected_name}


# ── Field Extraction ──────────────────────────────────────────────────────────

def extract_fields(full_text: str, ocr_data: dict) -> list[dict]:
    """
    Heuristic field extractor.
    Handles:
      - "Label: value" patterns
      - "Label ______ value" (underline-fill patterns)
      - Checkbox patterns (☐ / □ / [ ] / [x] / ✓)
      - Two-column table rows (label left, value right)
    Returns list of {"field": str, "value": str, "confidence": str}
    """
    fields = []
    seen   = set()

    lines = [l.strip() for l in full_text.splitlines() if l.strip()]

    for line in lines:
        # ── Checkbox lines ────────────────────────────────────────────
        checkbox_match = re.match(
            r"^[\[\(]?([xX✓✗]?)[\]\)]?\s*[☐☑□■]?\s*(.+)$", line
        )
        if re.search(r"[\[\(☐☑□■✓✗]", line):
            checked = bool(re.search(r"[xX✓✗☑■]", line[:4]))
            label   = re.sub(r"^[\[\(]?[xX✓✗☑■ ]?[\]\)]?\s*[☐☑□■]?\s*", "", line).strip()
            label   = re.sub(r"[_]{2,}", "", label).strip()
            if label and label.lower() not in seen:
                seen.add(label.lower())
                fields.append({
                    "field":      label,
                    "value":      "Yes" if checked else "No",
                    "type":       "checkbox",
                    "confidence": "medium"
                })
            continue

        # ── "Label: Value" pattern ────────────────────────────────────
        colon_match = re.match(r"^([^:]{2,60}):\s*(.*)$", line)
        if colon_match:
            label = colon_match.group(1).strip()
            value = colon_match.group(2).strip()
            # Strip trailing underscores from value (blank lines)
            value = re.sub(r"[_]{3,}$", "", value).strip()
            if label and label.lower() not in seen:
                seen.add(label.lower())
                fields.append({
                    "field":      label,
                    "value":      value,
                    "type":       "text",
                    "confidence": "high" if value else "medium"
                })
            continue

        # ── "Label ________" (blank fill line) ───────────────────────
        blank_match = re.match(r"^([A-Za-z][^_]{2,50}?)\s*[_]{3,}\s*(.*)$", line)
        if blank_match:
            label = blank_match.group(1).strip()
            value = blank_match.group(2).strip()
            if label and label.lower() not in seen:
                seen.add(label.lower())
                fields.append({
                    "field":      label,
                    "value":      value,
                    "type":       "text",
                    "confidence": "high" if value else "medium"
                })
            continue

    # ── Two-column layout: use bounding boxes ─────────────────────────
    fields += extract_two_column(ocr_data, seen)

    return fields


def extract_two_column(ocr_data: dict, seen: set) -> list[dict]:
    """
    Detect two-column label→value pairs by grouping words into rows
    and finding a column split midpoint.
    """
    results = []
    n       = len(ocr_data["text"])
    img_w   = 1  # relative

    # Group words by line_num + block_num
    rows: dict[tuple, list] = {}
    for i in range(n):
        conf = int(ocr_data["conf"][i]) if str(ocr_data["conf"][i]).lstrip("-").isdigit() else -1
        text = ocr_data["text"][i].strip()
        if conf < 30 or not text:
            continue
        key = (ocr_data["block_num"][i], ocr_data["line_num"][i])
        rows.setdefault(key, []).append({
            "text": text,
            "x":    ocr_data["left"][i],
            "y":    ocr_data["top"][i],
            "w":    ocr_data["width"][i],
            "conf": conf
        })

    if not rows:
        return results

    # Find approximate page width
    all_rights = [w["x"] + w["w"] for wlist in rows.values() for w in wlist]
    if all_rights:
        img_w = max(all_rights)

    midpoint = img_w * 0.45

    for words in rows.values():
        if len(words) < 2:
            continue
        words_sorted = sorted(words, key=lambda w: w["x"])
        left_words   = [w for w in words_sorted if w["x"] < midpoint]
        right_words  = [w for w in words_sorted if w["x"] >= midpoint]
        if not left_words or not right_words:
            continue

        label = " ".join(w["text"] for w in left_words).strip(": ")
        value = " ".join(w["text"] for w in right_words).strip()

        # Skip if looks like a sentence (>6 left words)
        if len(left_words) > 6:
            continue
        if label and label.lower() not in seen:
            seen.add(label.lower())
            results.append({
                "field":      label,
                "value":      value,
                "type":       "text",
                "confidence": "medium"
            })

    return results


# ── Auto-save ─────────────────────────────────────────────────────────────────

def auto_save_scan(uid: str, fields: list, language: str, original_filename: str):
    """
    Automatically save every scan as an Excel file the moment it's processed.
    Filename: YYYY-MM-DD_HH-MM-SS_<uid>.xlsx
    Includes a metadata sheet with scan info (time, language, original file).
    """
    from datetime import datetime
    try:
        timestamp   = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        save_name   = f"{timestamp}_{uid[:8]}.xlsx"
        save_path   = EXPORT_DIR / save_name

        wb = openpyxl.Workbook()

        # ── Sheet 1: Extracted fields ──────────────────────────────────
        ws = wb.active
        ws.title = "Extracted Fields"

        header_font  = Font(bold=True, color="FFFFFF", size=11)
        header_fill  = PatternFill("solid", fgColor="2E6B87")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        thin_border  = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        conf_colors = {"high": "D9F2E6", "medium": "FFF9C4", "low": "FFE0E0"}

        headers = ["Field", "Value", "Type", "Confidence"]
        ws.append(headers)
        for col in range(1, 5):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = center_align; cell.border = thin_border

        for i, f in enumerate(fields, 2):
            ws.cell(row=i, column=1).value = f.get("field", "")
            ws.cell(row=i, column=2).value = f.get("value", "")
            ws.cell(row=i, column=3).value = f.get("type", "text")
            ws.cell(row=i, column=4).value = f.get("confidence", "")
            color = conf_colors.get(f.get("confidence", ""), "FFFFFF")
            for col in range(1, 5):
                cell = ws.cell(row=i, column=col)
                cell.alignment = left_align if col <= 2 else center_align
                cell.fill   = PatternFill("solid", fgColor=color)
                cell.border = thin_border

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 14
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:D{len(fields) + 1}"

        # ── Sheet 2: Scan metadata ─────────────────────────────────────
        ws2 = wb.create_sheet("Scan Info")
        meta_rows = [
            ("Scanned at",        datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Original file",     original_filename),
            ("Detected language", language),
            ("Total fields",      len(fields)),
            ("High confidence",   sum(1 for f in fields if f.get("confidence") == "high")),
            ("Medium confidence", sum(1 for f in fields if f.get("confidence") == "medium")),
            ("Needs review",      sum(1 for f in fields if not f.get("value"))),
            ("Scan ID",           uid),
        ]
        for row in meta_rows:
            ws2.append(row)
        for row in ws2.iter_rows():
            for cell in row:
                cell.alignment = left_align
                cell.border    = thin_border
        ws2.column_dimensions["A"].width = 22
        ws2.column_dimensions["B"].width = 30

        wb.save(save_path)
        app.logger.info(f"Auto-saved scan: {save_path}")

    except Exception:
        app.logger.exception("Auto-save failed (non-fatal)")


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/process", methods=["POST"])
def process_form():
    """
    POST /api/process
    Accepts: multipart/form-data with field 'file'
    Returns: JSON { fields: [...], full_text: str, image_url: str }
    """
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files["file"]
    if not file or not file.filename:
        return jsonify({"error": "Empty file"}), 400
    if not allowed_file(file.filename):
        return jsonify({"error": f"File type not supported. Allowed: {', '.join(ALLOWED_EXTENSIONS)}"}), 400

    # Save upload
    uid      = uuid.uuid4().hex
    ext      = file.filename.rsplit(".", 1)[1].lower()
    filename = f"{uid}.{ext}"
    save_path = UPLOAD_DIR / filename
    file.save(save_path)

    try:
        # Handle PDF: convert first page to image
        if ext == "pdf":
            from pdf2image import convert_from_path
            pages = convert_from_path(save_path, dpi=200, first_page=1, last_page=1)
            if not pages:
                return jsonify({"error": "Could not read PDF"}), 400
            pil_img = pages[0]
        else:
            pil_img = Image.open(save_path)

        # Preprocess
        preprocessed = preprocess_image(pil_img)

        # OCR with auto language detection
        ocr_result = run_ocr(preprocessed)

        # Extract fields
        fields = extract_fields(ocr_result["full_text"], ocr_result["data"])

        # Save preprocessed image for preview
        preview_name = f"{uid}_preview.png"
        preview_path = UPLOAD_DIR / preview_name
        cv2.imwrite(str(preview_path), preprocessed)

        # ── Auto-save scan to exports folder immediately ──────────────
        detected_language = ocr_result.get("detected_language", "English")
        auto_save_scan(uid, fields, detected_language, file.filename)

        return jsonify({
            "uid":               uid,
            "fields":            fields,
            "full_text":         ocr_result["full_text"],
            "image_url":         f"/uploads/{preview_name}",
            "detected_language": detected_language
        })

    except Exception as e:
        app.logger.exception("Processing error")
        return jsonify({"error": str(e)}), 500


@app.route("/api/export", methods=["POST"])
def export_data():
    """
    POST /api/export
    Body: { format: "xlsx"|"csv", fields: [...], filename: str }
    Returns: file download
    """
    body     = request.get_json(force=True)
    fmt      = body.get("format", "xlsx")
    fields   = body.get("fields", [])
    form_name = body.get("filename", "form_export")

    if not fields:
        return jsonify({"error": "No fields to export"}), 400

    from datetime import datetime
    uid       = uuid.uuid4().hex[:8]
    safe_name = re.sub(r"[^\w\-]", "_", form_name)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    export_filename = f"{timestamp}_{safe_name}.{fmt}"

    if fmt == "csv":
        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=["field", "value", "type", "confidence"])
        writer.writeheader()
        for f in fields:
            writer.writerow({
                "field":      f.get("field", ""),
                "value":      f.get("value", ""),
                "type":       f.get("type", "text"),
                "confidence": f.get("confidence", "")
            })
        output.seek(0)
        csv_bytes = output.getvalue().encode("utf-8-sig")
        # Save a copy to exports folder
        (EXPORT_DIR / export_filename).write_bytes(csv_bytes)
        return send_file(
            BytesIO(csv_bytes),
            mimetype="text/csv",
            as_attachment=True,
            download_name=export_filename
        )

    else:  # xlsx
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Extracted Fields"

        # Styles
        header_font  = Font(bold=True, color="FFFFFF", size=11)
        header_fill  = PatternFill("solid", fgColor="2E6B87")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        thin_border  = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        conf_colors = {"high": "D9F2E6", "medium": "FFF9C4", "low": "FFE0E0"}

        # Header row
        headers = ["Field", "Value", "Type", "Confidence"]
        ws.append(headers)
        for col, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center_align
            cell.border    = thin_border

        # Data rows
        for i, f in enumerate(fields, 2):
            ws.cell(row=i, column=1).value = f.get("field", "")
            ws.cell(row=i, column=2).value = f.get("value", "")
            ws.cell(row=i, column=3).value = f.get("type", "text")
            ws.cell(row=i, column=4).value = f.get("confidence", "")

            conf  = f.get("confidence", "")
            color = conf_colors.get(conf, "FFFFFF")
            for col in range(1, 5):
                cell = ws.cell(row=i, column=col)
                cell.alignment = left_align if col <= 2 else center_align
                cell.fill      = PatternFill("solid", fgColor=color)
                cell.border    = thin_border

        # Column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 14

        # Freeze header
        ws.freeze_panes = "A2"

        # Auto-filter
        ws.auto_filter.ref = f"A1:D{len(fields) + 1}"

        # Summary sheet
        ws2 = wb.create_sheet("Summary")
        ws2.append(["Form Name", form_name])
        ws2.append(["Total Fields", len(fields)])
        ws2.append(["High Confidence",   sum(1 for f in fields if f.get("confidence") == "high")])
        ws2.append(["Medium Confidence", sum(1 for f in fields if f.get("confidence") == "medium")])
        ws2.append(["Low Confidence",    sum(1 for f in fields if f.get("confidence") == "low")])
        ws2.append(["Needs Review",      sum(1 for f in fields if not f.get("value"))])

        for row in ws2.iter_rows():
            for cell in row:
                cell.alignment = left_align
                cell.border    = thin_border
        ws2.column_dimensions["A"].width = 22
        ws2.column_dimensions["B"].width = 20

        out = BytesIO()
        wb.save(out)
        xlsx_bytes = out.getvalue()
        # Save a copy to exports folder
        (EXPORT_DIR / export_filename).write_bytes(xlsx_bytes)
        return send_file(
            BytesIO(xlsx_bytes),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=export_filename
        )


@app.route("/uploads/<path:filename>")
def serve_upload(filename):
    return send_from_directory(UPLOAD_DIR, filename)


@app.route("/exports")
def exports_page():
    """Admin page listing all saved exports."""
    from datetime import datetime
    files = sorted(EXPORT_DIR.glob("*.*"), key=lambda f: f.stat().st_mtime, reverse=True)
    export_list = []
    for f in files:
        stat = f.stat()
        # Auto-saved files follow pattern: YYYY-MM-DD_HH-MM-SS_<uid8>.xlsx
        import re as _re
        auto_saved = bool(_re.match(r"\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2}_[a-f0-9]{8}\.xlsx", f.name))
        export_list.append({
            "name":       f.name,
            "size_kb":    round(stat.st_size / 1024, 1),
            "modified":   datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
            "auto_saved": auto_saved,
        })
    return render_template("exports.html", exports=export_list)


@app.route("/exports/download/<path:filename>")
def download_export(filename):
    """Download a specific saved export."""
    safe = Path(filename).name  # prevent path traversal
    return send_from_directory(EXPORT_DIR, safe, as_attachment=True)


@app.route("/exports/delete/<path:filename>", methods=["POST"])
def delete_export(filename):
    """Delete a saved export."""
    safe = Path(filename).name
    target = EXPORT_DIR / safe
    if target.exists():
        target.unlink()
    return jsonify({"ok": True})


@app.route("/api/health")
def health():
    try:
        version = pytesseract.get_tesseract_version()
        return jsonify({"status": "ok", "tesseract": str(version)})
    except Exception as e:
        return jsonify({"status": "error", "detail": str(e)}), 500


# ── Entry ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5050))
    debug = os.environ.get("FLASK_DEBUG", "false").lower() == "true"
    print(f"\n✅  Form Digitizer running at http://localhost:{port}\n")
    app.run(host="0.0.0.0", port=port, debug=debug)
